"""
SPX 500 + NDX 100 每日筛选 + 邮件发送
条件：股价低于200日均线 20% 及以上
每天自动运行，结果Excel发送到指定邮箱

首次使用前请修改下面的 SENDER_PASSWORD
"""

import os
import time
import smtplib
from datetime import datetime
from email.message import EmailMessage

import yfinance as yf
import pandas as pd

# ═══════════════════════════════════════════════════
#  ⚙️  配置区域
#  在 GitHub Actions 上运行时，密码从 Secrets 读取；
#  本地运行时，改下面 LOCAL_PASSWORD 即可。
# ═══════════════════════════════════════════════════
SENDER_EMAIL    = os.environ.get("SENDER_EMAIL",    "croweelterry@gmail.com")
RECIPIENT_EMAIL = os.environ.get("RECIPIENT_EMAIL", "tangpingtaotpt@gmail.com")

LOCAL_PASSWORD  = "在这里填入16位App Password"   # ← 只在本地跑时才需要填
SENDER_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", LOCAL_PASSWORD)
# ═══════════════════════════════════════════════════

THRESHOLD = 0.20
OUTPUT_DIR = os.environ.get("OUTPUT_DIR", os.path.dirname(os.path.abspath(__file__)))

SPX_TICKERS = [
    "MMM","ABT","ABBV","ACN","ADBE","AMD","AES","AFL","A","APD","ABNB",
    "AKAM","ALB","ARE","ALGN","ALLE","LNT","ALL","GOOGL","GOOG","MO","AMZN",
    "AMCR","AEE","AEP","AXP","AIG","AMT","AWK","AMP","AME","AMGN",
    "APH","ADI","AAPL","AMAT","APTV","ACGL","ANET","AJG","AIZ","T","ATO",
    "ADSK","ADP","AZO","AVB","AVY","AXON","BKR","BALL","BAC","BBWI",
    "BAX","BDX","WRB","BBY","BIO","BIIB","BLK","BX","BA","BKNG","BWA",
    "BSX","BMY","AVGO","BR","BRO","BLDR","CHRW","CDNS","CPT","CPB","COF",
    "CAH","KMX","CCL","CARR","CAT","CBOE","CBRE","CDW","CE","COR","CNC",
    "CNP","CF","CRL","SCHW","CHTR","CVX","CMG","CB","CHD","CI","CINF",
    "CTAS","CSCO","C","CFG","CLX","CME","CMS","KO","CTSH","CL","CMCSA",
    "CAG","COP","ED","STZ","CEG","COO","CPRT","GLW","CPAY","CTVA","CSGP",
    "COST","CCI","CSX","CMI","CVS","DHI","DHR","DRI","DVA","DECK",
    "DE","DELL","DAL","DVN","DXCM","FANG","DLR","DG","DLTR","D",
    "DPZ","DOV","DOW","DTE","DUK","DD","EMN","ETN","EBAY","ECL","EIX",
    "EW","EA","ELV","LLY","EMR","ENPH","ETR","EOG","EPAM","EQT","EFX",
    "EQIX","EQR","ESS","EL","ETSY","EG","ES","EXC","EXPE","EXPD","EXR",
    "XOM","FFIV","FDS","FICO","FAST","FRT","FDX","FIS","FITB","FSLR","FE",
    "FMC","F","FTNT","FTV","FOXA","FOX","BEN","FCX","GRMN","IT","GE",
    "GEHC","GEV","GEN","GNRC","GD","GIS","GM","GPC","GILD","GPN","GL",
    "GDDY","GS","HAL","HIG","HAS","HCA","DOC","HSIC","HSY","HPE",
    "HLT","HOLX","HD","HON","HRL","HST","HWM","HPQ","HUBB","HUM","HBAN",
    "HII","IBM","IEX","IDXX","ITW","INCY","IR","PODD","INTC","ICE","IFF",
    "IP","INTU","ISRG","IVZ","INVH","IQV","IRM","J","JBL","JCI","JPM",
    "KVUE","KDP","KEY","KEYS","KMB","KIM","KMI","KKR","KLAC","KHC",
    "KR","LHX","LH","LRCX","LW","LVS","LDOS","LEN","LIN","LYV","LKQ",
    "LMT","L","LOW","LULU","LYB","MTB","MPC","MKTX","MAR","MLM","MAS","MA",
    "MTCH","MCD","MCK","MDT","MRK","META","MET","MTD",
    "MGM","MCHP","MU","MSFT","MAA","MRNA","MHK","MOH","TAP","MDLZ","MPWR",
    "MNST","MCO","MS","MOS","MSI","MSCI","NDAQ","NTAP","NFLX","NEM","NWSA",
    "NWS","NEE","NKE","NI","NDSN","NSC","NTRS","NOC","NCLH","NRG","NUE",
    "NVDA","NVR","NXPI","ORLY","OXY","ODFL","OMC","ON","OKE","ORCL","OTIS",
    "OGN","PCAR","PKG","PLTR","PH","PAYX","PAYC","PYPL","PNR","PEP","PFE",
    "PCG","PM","PSX","PNW","PNC","POOL","PPG","PPL","PFG","PG","PGR","PLD",
    "PRU","PEG","PTC","PSA","PHM","QRVO","PWR","QCOM","DGX","RL","RJF",
    "RTX","O","REG","REGN","RF","RSG","RMD","RVTY","ROK","ROL","ROP","ROST",
    "RCL","SPGI","CRM","SBAC","SLB","STX","SRE","NOW","SHW","SPG","SWKS",
    "SJM","SW","SNA","SOLV","SO","LUV","SWK","SBUX","STT","STLD","STE",
    "SYK","SYF","SNPS","SYY","TMUS","TROW","TTWO","TPR","TRGP","TGT","TEL",
    "TDY","TFX","TER","TSLA","TXN","TXT","TMO","TJX","TSCO","TT","TDG",
    "TRV","TRMB","TFC","TYL","TSN","USB","UBER","UDR","ULTA","UNP","UAL",
    "UPS","URI","UNH","UHS","VLO","VTR","VRSN","VRSK","VZ","VRTX","VLTO",
    "VMC","WMT","DIS","WM","WAT","WEC","WFC","WELL","WST","WDC","WY",
    "WHR","WMB","WTW","GWW","WYNN","XEL","XYL","YUM","ZBRA","ZBH","ZTS"
]

NDX_TICKERS = [
    "ADBE","ADP","ABNB","ALGN","GOOGL","GOOG","AMZN","AMD","AEP","AMGN",
    "ADI","AAPL","AMAT","ASML","TEAM","ADSK","BIIB","BKNG","AVGO","CDNS",
    "CHTR","CTAS","CSCO","CTSH","CMCSA","CEG","CPRT","CSGP","COST","CRWD",
    "CSX","DDOG","DXCM","FANG","DLTR","EA","EXC","FAST","FTNT","GILD",
    "HON","IDXX","INTC","INTU","ISRG","KDP","KLAC","LRCX","LIN","MAR",
    "MRVL","META","MCHP","MU","MSFT","MRNA","MNST","NFLX","NVDA","NXPI",
    "ORLY","ON","PCAR","PANW","PAYX","PYPL","PEP","QCOM","REGN","ROST",
    "CRM","SNPS","SBUX","TMUS","TSLA","TXN","TTWO","VRSK","VRTX","WDAY","ZS"
]

# 道琼斯综合平均指数 (DJCA) = 工业30 + 运输20 + 公用事业15 = 65只

# 道琼斯工业平均指数 (DJIA) — 30只
DJIA_TICKERS = [
    "AAPL","AMGN","AXP","BA","CAT","CRM","CSCO","CVX","DIS","DOW",
    "GS","HD","HON","IBM","JNJ","JPM","KO","MCD","MMM","MRK",
    "MSFT","NKE","PG","SHW","TRV","UNH","V","VZ","WMT","WBA"
]

# 道琼斯运输业平均指数 (DJTA) — 20只
DJTA_TICKERS = [
    "ALK","AAL","CAR","CHRW","CSX","DAL","EXPD","FDX","JBHT","JBLU",
    "KEX","LSTR","MATX","NSC","ODFL","R","UAL","UNP","UPS","XPO"
]

# 道琼斯公用事业平均指数 (DJUA) — 15只
DJUA_TICKERS = [
    "AEE","AEP","ATO","CNP","D","ED","EIX","ES","EXC","FE",
    "NEE","PEG","PPL","SO","SRE"
]

# 合并为道琼斯综合 (DJCA)
DJI_TICKERS = list(dict.fromkeys(DJIA_TICKERS + DJTA_TICKERS + DJUA_TICKERS))

COMPANY_NAMES = {
    "MMM":"3M","ABT":"Abbott Labs","ABBV":"AbbVie","ACN":"Accenture","ADBE":"Adobe",
    "AMD":"AMD","AES":"AES Corp","AFL":"Aflac","A":"Agilent","APD":"Air Products",
    "ABNB":"Airbnb","AKAM":"Akamai","ALB":"Albemarle","ARE":"Alexandria RE",
    "ALGN":"Align Technology","ALL":"Allstate","GOOGL":"Alphabet A","GOOG":"Alphabet C",
    "MO":"Altria","AMZN":"Amazon","AEP":"AEP","AXP":"American Express","AIG":"AIG",
    "AMT":"American Tower","AWK":"American Water","AMP":"Ameriprise","AMGN":"Amgen",
    "APH":"Amphenol","ADI":"Analog Devices","AAPL":"Apple","AMAT":"Applied Materials",
    "ANET":"Arista Networks","T":"AT&T","ADSK":"Autodesk","ADP":"ADP","AZO":"AutoZone",
    "BKR":"Baker Hughes","BAC":"Bank of America","BAX":"Baxter",
    "BDX":"Becton Dickinson","BBY":"Best Buy","BIIB":"Biogen","BLK":"BlackRock",
    "BX":"Blackstone","BA":"Boeing","BKNG":"Booking Holdings","BSX":"Boston Scientific",
    "BMY":"Bristol-Myers","AVGO":"Broadcom","BRO":"Brown & Brown","BLDR":"Builders FirstSource",
    "CDNS":"Cadence Design","COF":"Capital One","CAH":"Cardinal Health","KMX":"CarMax",
    "CCL":"Carnival","CARR":"Carrier Global","CAT":"Caterpillar","CBOE":"Cboe Global",
    "CBRE":"CBRE Group","CDW":"CDW Corp","COR":"Cencora","CNC":"Centene",
    "SCHW":"Charles Schwab","CHTR":"Charter Comm","CVX":"Chevron","CMG":"Chipotle",
    "CB":"Chubb","CI":"Cigna","CTAS":"Cintas","CSCO":"Cisco","C":"Citigroup",
    "CLX":"Clorox","CME":"CME Group","KO":"Coca-Cola","CTSH":"Cognizant",
    "CL":"Colgate-Palmolive","CMCSA":"Comcast","CAG":"Conagra","COP":"ConocoPhillips",
    "ED":"Consolidated Edison","STZ":"Constellation Brands","CEG":"Constellation Energy",
    "CPRT":"Copart","GLW":"Corning","CTVA":"Corteva","CSGP":"CoStar Group",
    "COST":"Costco","CCI":"Crown Castle","CSX":"CSX Corp",
    "CMI":"Cummins","CVS":"CVS Health","DHI":"D.R. Horton","DHR":"Danaher",
    "DRI":"Darden Restaurants","DE":"Deere & Co","DELL":"Dell Technologies",
    "DAL":"Delta Air Lines","DVN":"Devon Energy","DXCM":"DexCom","FANG":"Diamondback Energy",
    "DLR":"Digital Realty","DG":"Dollar General",
    "DLTR":"Dollar Tree","D":"Dominion Energy","DOV":"Dover Corp","DOW":"Dow Inc",
    "DUK":"Duke Energy","DD":"DuPont","ETN":"Eaton","EBAY":"eBay","ECL":"Ecolab",
    "EW":"Edwards Lifesciences","EA":"Electronic Arts","ELV":"Elevance Health",
    "LLY":"Eli Lilly","EMR":"Emerson Electric","ENPH":"Enphase Energy","EOG":"EOG Resources",
    "EQT":"EQT Corp","EFX":"Equifax","EQIX":"Equinix","EL":"Estee Lauder",
    "ETSY":"Etsy","EXC":"Exelon","EXPE":"Expedia","XOM":"ExxonMobil",
    "FDS":"FactSet","FICO":"Fair Isaac","FAST":"Fastenal","FDX":"FedEx",
    "FIS":"Fidelity National Info","FITB":"Fifth Third Bancorp","FSLR":"First Solar",
    "FMC":"FMC Corp","F":"Ford Motor","FTNT":"Fortinet",
    "FCX":"Freeport-McMoRan","GRMN":"Garmin","IT":"Gartner","GE":"GE Aerospace",
    "GEHC":"GE HealthCare","GEV":"GE Vernova","GD":"General Dynamics",
    "GIS":"General Mills","GM":"General Motors","GPC":"Genuine Parts",
    "GILD":"Gilead Sciences","GPN":"Global Payments","GS":"Goldman Sachs",
    "HAL":"Halliburton","HIG":"Hartford Financial","HAS":"Hasbro","HCA":"HCA Healthcare",
    "HSY":"Hershey","HPE":"HP Enterprise","HLT":"Hilton","HOLX":"Hologic",
    "HD":"Home Depot","HON":"Honeywell","HRL":"Hormel Foods","HPQ":"HP Inc",
    "HUBB":"Hubbell","HUM":"Humana","HBAN":"Huntington Bancshares","IBM":"IBM",
    "IDXX":"IDEXX Labs","ITW":"Illinois Tool Works","INTC":"Intel","ICE":"Intercontinental Exchange",
    "INTU":"Intuit","ISRG":"Intuitive Surgical","IQV":"IQVIA","IRM":"Iron Mountain",
    "J":"Jacobs Solutions","JBL":"Jabil","JCI":"Johnson Controls","JPM":"JPMorgan Chase",
    "KDP":"Keurig Dr Pepper","KEY":"KeyCorp","KEYS":"Keysight","KMB":"Kimberly-Clark",
    "KIM":"Kimco Realty","KMI":"Kinder Morgan","KKR":"KKR & Co","KLAC":"KLA Corp",
    "KHC":"Kraft Heinz","KR":"Kroger","LHX":"L3Harris","LH":"LabCorp",
    "LRCX":"Lam Research","LVS":"Las Vegas Sands","LDOS":"Leidos","LEN":"Lennar",
    "LIN":"Linde","LMT":"Lockheed Martin","LOW":"Lowe's","LULU":"Lululemon",
    "LYB":"LyondellBasell","MTB":"M&T Bank","MPC":"Marathon Petroleum","MAR":"Marriott",
    "MLM":"Martin Marietta","MAS":"Masco","MA":"Mastercard","MCD":"McDonald's",
    "MCK":"McKesson","MDT":"Medtronic","MRK":"Merck","META":"Meta Platforms",
    "MET":"MetLife","MTD":"Mettler-Toledo","MGM":"MGM Resorts","MCHP":"Microchip Technology",
    "MU":"Micron Technology","MSFT":"Microsoft","MRNA":"Moderna","MOH":"Molina Healthcare",
    "TAP":"Molson Coors","MDLZ":"Mondelez","MPWR":"Monolithic Power","MNST":"Monster Beverage",
    "MCO":"Moody's","MS":"Morgan Stanley","MSI":"Motorola Solutions","MSCI":"MSCI Inc",
    "NDAQ":"Nasdaq Inc","NTAP":"NetApp","NFLX":"Netflix","NEM":"Newmont",
    "NEE":"NextEra Energy","NKE":"Nike","NSC":"Norfolk Southern","NOC":"Northrop Grumman",
    "NRG":"NRG Energy","NUE":"Nucor","NVDA":"Nvidia","NXPI":"NXP Semiconductors",
    "ORLY":"O'Reilly Auto","OXY":"Occidental Petroleum","ODFL":"Old Dominion Freight",
    "ON":"ON Semiconductor","OKE":"ONEOK","ORCL":"Oracle","OTIS":"Otis Worldwide",
    "PCAR":"PACCAR","PKG":"Packaging Corp","PLTR":"Palantir","PH":"Parker Hannifin",
    "PAYX":"Paychex","PYPL":"PayPal","PEP":"PepsiCo","PFE":"Pfizer","PCG":"PG&E",
    "PM":"Philip Morris","PSX":"Phillips 66","PNC":"PNC Financial","PPG":"PPG Industries",
    "PG":"Procter & Gamble","PGR":"Progressive","PLD":"Prologis","PRU":"Prudential Financial",
    "PEG":"PSEG","QCOM":"Qualcomm","RL":"Ralph Lauren","RTX":"RTX Corp",
    "O":"Realty Income","REGN":"Regeneron","RF":"Regions Financial","RSG":"Republic Services",
    "RMD":"ResMed","ROK":"Rockwell Automation","ROP":"Roper Technologies","ROST":"Ross Stores",
    "RCL":"Royal Caribbean","SPGI":"S&P Global","CRM":"Salesforce","SLB":"SLB",
    "STX":"Seagate","SRE":"Sempra","NOW":"ServiceNow","SHW":"Sherwin-Williams",
    "SPG":"Simon Property","SJM":"J.M. Smucker","SNA":"Snap-on","SO":"Southern Company",
    "LUV":"Southwest Airlines","SWK":"Stanley Black & Decker","SBUX":"Starbucks",
    "STT":"State Street","STLD":"Steel Dynamics","STE":"Steris","SYK":"Stryker",
    "SYF":"Synchrony Financial","SNPS":"Synopsys","SYY":"Sysco","TMUS":"T-Mobile US",
    "TROW":"T. Rowe Price","TTWO":"Take-Two Interactive","TGT":"Target",
    "TEL":"TE Connectivity","TER":"Teradyne","TSLA":"Tesla","TXN":"Texas Instruments",
    "TMO":"Thermo Fisher","TJX":"TJX Companies","TSCO":"Tractor Supply",
    "TT":"Trane Technologies","TDG":"TransDigm","TRV":"Travelers","TFC":"Truist Financial",
    "TSN":"Tyson Foods","USB":"US Bancorp","UBER":"Uber","ULTA":"Ulta Beauty",
    "UNP":"Union Pacific","UAL":"United Airlines","UPS":"UPS","URI":"United Rentals",
    "UNH":"UnitedHealth","VLO":"Valero Energy","VZ":"Verizon","VRTX":"Vertex Pharmaceuticals",
    "VMC":"Vulcan Materials","WMT":"Walmart","DIS":"Walt Disney","WM":"Waste Management",
    "WEC":"WEC Energy","WFC":"Wells Fargo","WELL":"Welltower","WDC":"Western Digital",
    "WY":"Weyerhaeuser","WMB":"Williams Companies","GWW":"WW Grainger","WYNN":"Wynn Resorts",
    "XEL":"Xcel Energy","XYL":"Xylem","YUM":"Yum! Brands","ZBRA":"Zebra Technologies",
    "ZBH":"Zimmer Biomet","ZTS":"Zoetis","ASML":"ASML Holding","TEAM":"Atlassian",
    "CRWD":"CrowdStrike","DDOG":"Datadog","MRVL":"Marvell Technology",
    "PANW":"Palo Alto Networks","WDAY":"Workday","ZS":"Zscaler","VRSK":"Verisk Analytics",
    "JNJ":"Johnson & Johnson","V":"Visa","WBA":"Walgreens Boots Alliance",
    # 道琼斯运输业 (DJTA)
    "ALK":"Alaska Air Group","AAL":"American Airlines","CAR":"Avis Budget Group",
    "CHRW":"C.H. Robinson","JBHT":"J.B. Hunt Transport","JBLU":"JetBlue Airways",
    "KEX":"Kirby Corp","LSTR":"Landstar System","MATX":"Matson Inc",
    "R":"Ryder System","XPO":"XPO Inc",
    # 道琼斯公用事业 (DJUA)
    "AEE":"Ameren","ATO":"Atmos Energy","CNP":"CenterPoint Energy",
    "EIX":"Edison International","ES":"Eversource Energy","FE":"FirstEnergy",
}


# ═══════════════════════════════════════════════════
#  罗素1000 成分股（运行时从 Wikipedia 动态获取）
# ═══════════════════════════════════════════════════
R1K_CACHE_FILE = os.path.join(OUTPUT_DIR, "russell1000_cache.csv")
R1K_CACHE_AGE_DAYS = None   # 用了缓存时记录其新鲜度


def _clean_symbol(s):
    """维基/iShares 用 . 表示股份类别，Yahoo 用 -（BRK.B -> BRK-B）"""
    import re
    s = str(s).strip().upper()
    if not re.fullmatch(r"[A-Z][A-Z.\-]{0,6}", s):
        return None
    return s.replace(".", "-")


def _r1k_from_ishares():
    """来源一：iShares IWB ETF 官方持仓 CSV。最稳定，含公司名。"""
    import requests, io, csv, re

    urls = [
        ("https://www.ishares.com/us/products/239707/"
         "ishares-russell-1000-etf/1467271812596.ajax"
         "?fileType=csv&fileName=IWB_holdings&dataType=fund"),
        ("https://www.ishares.com/us/products/239707/"
         "ishares-russell-1000-etf/1521942788811.ajax"
         "?fileType=csv&fileName=IWB_holdings&dataType=fund"),
    ]

    last_err = None
    for url in urls:
        try:
            r = requests.get(url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "Accept": "text/csv,*/*",
            }, timeout=45)
            r.raise_for_status()
            text = r.text
            lines = [l for l in text.splitlines() if l.strip()]

            # 宽松找表头：某一行同时含 Ticker 和 Name
            hdr = None
            for i, l in enumerate(lines[:60]):
                low = l.lower()
                if "ticker" in low and "name" in low:
                    hdr = i
                    break
            if hdr is None:
                last_err = f"未找到表头（前3行示例：{lines[:3]}）"
                continue

            reader = csv.DictReader(io.StringIO("\n".join(lines[hdr:])))
            # 找出真正叫 Ticker / Name / Asset Class 的列（可能有空格或引号）
            fns = reader.fieldnames or []
            col_tk = next((c for c in fns if c and c.strip().lower() == "ticker"), None)
            col_nm = next((c for c in fns if c and c.strip().lower() == "name"), None)
            col_ac = next((c for c in fns if c and "asset class" in c.strip().lower()), None)
            if col_tk is None:
                last_err = f"表头里没有 Ticker 列，实际列名：{fns[:8]}"
                continue

            tickers, names = [], {}
            for row in reader:
                sym = _clean_symbol(row.get(col_tk, ""))
                if not sym:
                    continue
                if col_ac:
                    ac = (row.get(col_ac) or "").strip().lower()
                    if ac and ac != "equity":
                        continue
                tickers.append(sym)
                if col_nm:
                    nm = (row.get(col_nm) or "").strip()
                    if nm:
                        names[sym] = nm

            tickers = list(dict.fromkeys(tickers))
            if len(tickers) >= 500:
                return tickers, names
            last_err = f"只解析出 {len(tickers)} 只"
        except Exception as e:
            last_err = f"{type(e).__name__}: {e}"

    raise RuntimeError(last_err or "未知错误")


def _r1k_from_wikipedia():
    """昨天实测可用的原版：解析维基表格里的 Symbol 列。"""
    import re, requests
    from io import StringIO

    url = "https://en.wikipedia.org/wiki/Russell_1000_Index"
    html = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30).text
    tables = pd.read_html(StringIO(html))
    print(f"    页面共解析出 {len(tables)} 张表")

    best_tickers, best_names = [], {}
    for i, t in enumerate(tables):
        cols = [str(c) for c in t.columns]
        sym_col = next((c for c in cols if re.search(r"symbol|ticker", c, re.I)), None)
        if sym_col is None:
            if t.shape[0] >= 100:      # 大表但没symbol列，值得报告
                print(f"    表#{i}: {t.shape[0]}行，列名={cols[:6]}（无Symbol列）")
            continue
        name_col = next((c for c in cols if re.search(r"company|name|security", c, re.I)), None)

        tickers, names = [], {}
        for _, row in t.iterrows():
            sym = str(row[sym_col]).strip()
            if not re.fullmatch(r"[A-Z][A-Z.\-]{0,6}", sym):
                continue
            yf_sym = sym.replace(".", "-")
            tickers.append(yf_sym)
            if name_col is not None:
                nm = re.sub(r"\[.*?\]", "", str(row[name_col])).strip()
                if nm and nm.lower() != "nan":
                    names[yf_sym] = nm

        print(f"    表#{i}: {t.shape[0]}行，Symbol列='{sym_col}'，提取到 {len(tickers)} 只")
        if len(tickers) > len(best_tickers):
            best_tickers, best_names = tickers, names

    return list(dict.fromkeys(best_tickers)), best_names


def _r1k_from_wiki_links():
    """备用：从页面里的个股链接提取代码。"""
    import re, requests
    html = requests.get(
        "https://en.wikipedia.org/wiki/Russell_1000_Index",
        headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
        timeout=45).text

    raw = []
    raw += re.findall(r"nasdaq\.com/market-activity/stocks/([A-Za-z][A-Za-z.\-]{0,6})", html)
    raw += re.findall(r"nyse\.com/quote/XNYS[:%3A]+([A-Za-z][A-Za-z.\-]{0,6})", html, re.I)
    tickers = [s for s in (_clean_symbol(r) for r in raw) if s]
    return list(dict.fromkeys(tickers)), {}


def _r1k_from_ishares():
    """来源一：iShares IWB ETF 官方持仓 CSV。最稳定，含公司名。"""
    import requests, io, csv, re

    urls = [
        ("https://www.ishares.com/us/products/239707/"
         "ishares-russell-1000-etf/1467271812596.ajax"
         "?fileType=csv&fileName=IWB_holdings&dataType=fund"),
        ("https://www.ishares.com/us/products/239707/"
         "ishares-russell-1000-etf/1521942788811.ajax"
         "?fileType=csv&fileName=IWB_holdings&dataType=fund"),
    ]

    last_err = None
    for url in urls:
        try:
            r = requests.get(url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
                "Accept": "text/csv,*/*",
            }, timeout=45)
            r.raise_for_status()
            text = r.text
            lines = [l for l in text.splitlines() if l.strip()]

            # 宽松找表头：某一行同时含 Ticker 和 Name
            hdr = None
            for i, l in enumerate(lines[:60]):
                low = l.lower()
                if "ticker" in low and "name" in low:
                    hdr = i
                    break
            if hdr is None:
                last_err = f"未找到表头（前3行示例：{lines[:3]}）"
                continue

            reader = csv.DictReader(io.StringIO("\n".join(lines[hdr:])))
            # 找出真正叫 Ticker / Name / Asset Class 的列（可能有空格或引号）
            fns = reader.fieldnames or []
            col_tk = next((c for c in fns if c and c.strip().lower() == "ticker"), None)
            col_nm = next((c for c in fns if c and c.strip().lower() == "name"), None)
            col_ac = next((c for c in fns if c and "asset class" in c.strip().lower()), None)
            if col_tk is None:
                last_err = f"表头里没有 Ticker 列，实际列名：{fns[:8]}"
                continue

            tickers, names = [], {}
            for row in reader:
                sym = _clean_symbol(row.get(col_tk, ""))
                if not sym:
                    continue
                if col_ac:
                    ac = (row.get(col_ac) or "").strip().lower()
                    if ac and ac != "equity":
                        continue
                tickers.append(sym)
                if col_nm:
                    nm = (row.get(col_nm) or "").strip()
                    if nm:
                        names[sym] = nm

            tickers = list(dict.fromkeys(tickers))
            if len(tickers) >= 500:
                return tickers, names
            last_err = f"只解析出 {len(tickers)} 只"
        except Exception as e:
            last_err = f"{type(e).__name__}: {e}"

    raise RuntimeError(last_err or "未知错误")


def _r1k_from_wikipedia():
    """来源：维基百科 Russell 1000 页面。

    该页表格只有一个 Company 列，公司名/交易所/代码挤在一格里，
    没有独立的 Symbol 列。所以不解析表格，直接从页面里的
    个股链接提取代码，形式如：
      nasdaq.com/market-activity/stocks/nvda   -> NVDA
      nyse.com/quote/XNYS:BRK.B                -> BRK-B
    """
    import requests, re
    html = requests.get(
        "https://en.wikipedia.org/wiki/Russell_1000_Index",
        headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
        timeout=45).text

    raw = []
    raw += re.findall(r"nasdaq\.com/market-activity/stocks/([A-Za-z][A-Za-z.\-]{0,6})", html)
    raw += re.findall(r"nyse\.com/quote/XNYS[:%3A]+([A-Za-z][A-Za-z.\-]{0,6})", html, re.I)
    raw += re.findall(r"cboe\.com/[^\"\'<>]*?/([A-Za-z][A-Za-z.\-]{0,6})/?[\"\'<]", html)

    tickers = []
    for r in raw:
        sym = _clean_symbol(r)
        if sym:
            tickers.append(sym)
    tickers = list(dict.fromkeys(tickers))

    if len(tickers) >= 500:
        return tickers, {}

    # 备用：万一以后页面加回了 Symbol 列，再试解析表格
    from io import StringIO
    try:
        tables = pd.read_html(StringIO(html))
    except Exception:
        return tickers, {}

    best_t = tickers
    for t in tables:
        cols = [str(c) for c in t.columns]
        sym_col = next((c for c in cols if re.search(r"symbol|ticker", c, re.I)), None)
        if sym_col is None:
            continue
        # 按整列取值，避免维基多列布局导致列名重复时取到 Series
        col = t[sym_col]
        if hasattr(col, "columns"):          # 同名列被取成 DataFrame
            col = col.iloc[:, 0]
        vals = col.dropna().astype(str).str.strip().tolist()
        tk = [s for s in (_clean_symbol(v) for v in vals) if s]
        if len(tk) > len(best_t):
            best_t = tk

    # 再兜一层：整页扫所有像股票代码的独立词
    if len(best_t) < 500:
        loose = re.findall(r">([A-Z]{1,5}(?:\.[A-Z])?)<", html)
        loose = [s for s in (_clean_symbol(v) for v in loose) if s]
        if len(set(loose)) > len(best_t):
            best_t = loose

    return list(dict.fromkeys(best_t)), {}


def _r1k_from_stockanalysis():
    """来源三：stockanalysis.com 的 IWB 持仓页，返回 JSON。"""
    import requests
    url = "https://stockanalysis.com/api/symbol/e/IWB/holdings"
    r = requests.get(url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Accept": "application/json",
    }, timeout=45)
    r.raise_for_status()
    payload = r.json()

    # 兼容几种可能的 JSON 结构
    rows = payload
    for key in ("data", "list", "holdings"):
        if isinstance(rows, dict) and key in rows:
            rows = rows[key]
    if isinstance(rows, dict):
        rows = next((v for v in rows.values() if isinstance(v, list)), [])

    tickers, names = [], {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        raw = row.get("s") or row.get("symbol") or row.get("ticker") or ""
        sym = _clean_symbol(raw)
        if not sym:
            continue
        tickers.append(sym)
        nm = (row.get("n") or row.get("name") or "").strip()
        if nm:
            names[sym] = nm
    return list(dict.fromkeys(tickers)), names


def _r1k_from_cache():
    """兜底来源：仓库里的缓存文件。同时记录缓存生成日期。"""
    global R1K_CACHE_AGE_DAYS
    df = pd.read_csv(R1K_CACHE_FILE)
    tickers = df["ticker"].dropna().astype(str).tolist()
    names = {}
    if "name" in df.columns:
        names = {r["ticker"]: r["name"] for _, r in df.dropna(subset=["name"]).iterrows()}

    if "fetched" in df.columns and len(df) > 0:
        try:
            d = pd.to_datetime(df["fetched"].iloc[0]).date()
            R1K_CACHE_AGE_DAYS = (datetime.now().date() - d).days
            print(f"    缓存生成于 {d}（{R1K_CACHE_AGE_DAYS} 天前）")
        except Exception:
            pass
    return tickers, names


def _save_r1k_cache(tickers, names):
    try:
        pd.DataFrame({
            "ticker": tickers,
            "name": [names.get(t, "") for t in tickers],
            "fetched": datetime.now().strftime("%Y-%m-%d"),
        }).to_csv(R1K_CACHE_FILE, index=False)
    except Exception:
        pass


def get_russell1000():
    """依次尝试多个来源，返回 (tickers, {ticker: 公司名})。"""
    sources = [
        ("本地缓存(iShares官方)",  _r1k_from_cache),
        ("维基百科(表格)",         _r1k_from_wikipedia),
        ("维基百科(链接)",         _r1k_from_wiki_links),
        ("iShares 在线接口",       _r1k_from_ishares),
        ("StockAnalysis IWB",      _r1k_from_stockanalysis),
    ]
    for label, fn in sources:
        try:
            tickers, names = fn()
            if len(tickers) >= 500:
                print(f"  ✓ 罗素1000 来源：{label} — {len(tickers)} 只")
                if not label.startswith("本地缓存"):
                    _save_r1k_cache(tickers, names)
                return tickers, names
            print(f"  ✗ {label}：只拿到 {len(tickers)} 只，数量不足，尝试下一个来源")
        except Exception as e:
            print(f"  ✗ {label} 失败：{type(e).__name__} {e}")

    print("  ⚠ 所有来源均失败，本次仅扫描 SPX/NDX/DJCA（覆盖范围会明显变小）")
    return [], {}


def fill_missing_names(tickers):
    """对仍缺公司名的股票，逐只从 Yahoo 补齐。只对筛选结果调用，数量少。"""
    found = {}
    for i, t in enumerate(tickers):
        try:
            info = yf.Ticker(t).info
            nm = info.get("shortName") or info.get("longName") or ""
            if nm:
                found[t] = nm.strip()
        except Exception:
            pass
        if (i + 1) % 20 == 0:
            print(f"    补全公司名 {i+1}/{len(tickers)} ...")
        time.sleep(0.1)
    return found


def batch_download_closes(tickers, batch_size=60):
    """批量下载收盘价，返回 {ticker: Close Series}。比逐只下载快十倍以上。"""
    out = {}
    total_batches = (len(tickers) + batch_size - 1) // batch_size

    for b in range(total_batches):
        chunk = tickers[b * batch_size:(b + 1) * batch_size]
        try:
            data = yf.download(
                chunk, period="300d", auto_adjust=True,
                progress=False, threads=True, group_by="column",
            )
            if data is None or data.empty:
                continue

            if isinstance(data.columns, pd.MultiIndex):
                if "Close" not in data.columns.get_level_values(0):
                    continue
                closes = data["Close"]
                for t in closes.columns:
                    s = closes[t].dropna()
                    if len(s) >= 30:
                        out[t] = s
            else:
                # 单只股票的情形
                if "Close" in data.columns and len(chunk) == 1:
                    s = data["Close"].dropna()
                    if len(s) >= 30:
                        out[chunk[0]] = s
        except Exception:
            pass

        print(f"  下载中 {min((b+1)*batch_size, len(tickers))}/{len(tickers)} ...")
        time.sleep(0.3)

    return out


def get_close_series(ticker):
    df = yf.download(ticker, period="300d", auto_adjust=True, progress=False)
    if df.empty:
        return None
    if isinstance(df.columns, pd.MultiIndex):
        cols = [c for c in df.columns if c[0] == "Close"]
        return df[cols[0]].dropna() if cols else None
    return df["Close"].dropna() if "Close" in df.columns else None


def run_screener():
    """跑筛选，返回 (DataFrame, 生成的Excel路径, 数据日期)"""
    spx_set = set(SPX_TICKERS)
    ndx_set = set(NDX_TICKERS)
    dji_set = set(DJI_TICKERS)

    print("  正在获取罗素1000成分股...")
    r1k, r1k_names = get_russell1000()
    r1k_set = set(r1k)
    print(f"  罗素1000：{len(r1k)} 只（含 {len(r1k_names)} 个公司名）")

    # 合并公司名：已有的中文名优先，维基的英文名补空缺
    names = dict(r1k_names)
    names.update(COMPANY_NAMES)

    all_tickers = list(dict.fromkeys(
        SPX_TICKERS + NDX_TICKERS + DJI_TICKERS + r1k
    ))
    print(f"  合并去重后共 {len(all_tickers)} 只，开始批量下载...\n")

    series_map = batch_download_closes(all_tickers)
    print(f"\n  成功获取 {len(series_map)} 只的价格数据，开始计算...\n")

    results = []
    data_date = None

    for ticker, series in series_map.items():
        try:
            current = float(series.iloc[-1])
            # 当日涨跌幅 = (今收 - 昨收) / 昨收
            day_chg = None
            if len(series) >= 2:
                prev = float(series.iloc[-2])
                if prev:
                    day_chg = round((current - prev) / prev * 100, 2)
            if data_date is None:
                data_date = series.index[-1].strftime("%Y-%m-%d")

            ma_days = min(200, len(series))
            ma200 = float(series.rolling(ma_days).mean().iloc[-1])
            if pd.isna(ma200) or ma200 == 0:
                continue

            deviation = (current - ma200) / ma200

            if deviation <= -THRESHOLD:
                idx = []
                if ticker in spx_set: idx.append("SPX")
                if ticker in ndx_set: idx.append("NDX")
                if ticker in dji_set: idx.append("DJI")
                if ticker in r1k_set: idx.append("R1K")
                results.append({
                    "股票代码": ticker,
                    "公司名称": names.get(ticker, ""),
                    "所属指数": "+".join(idx),
                    "收盘价(USD)": round(current, 2),
                    "当日涨跌%": day_chg,
                    "MA200(USD)": round(ma200, 2),
                    "偏离年线%": round(deviation * 100, 2),
                })
        except Exception:
            continue

    if not results:
        return None, None, data_date, len(r1k)

    df = pd.DataFrame(results).sort_values("偏离年线%")

    # 对仍缺公司名的（数量很少）去 Yahoo 补齐
    missing = df.loc[df["公司名称"] == "", "股票代码"].tolist()
    if missing:
        print(f"\n  有 {len(missing)} 只缺公司名，正在从 Yahoo 补全...")
        extra = fill_missing_names(missing)
        df["公司名称"] = df.apply(
            lambda r: extra.get(r["股票代码"], r["公司名称"]), axis=1
        )

    if data_date is None:
        data_date = datetime.now().strftime("%Y-%m-%d")
    fname = os.path.join(OUTPUT_DIR, f"SPX_NDX_DJCA_R1000_偏离年线20%以上_{data_date.replace('-','')}.xlsx")

    with pd.ExcelWriter(fname, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="低于年线20%以上", index=False)

    # 格式 + 筛选
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = load_workbook(fname)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    widths = {"股票代码":10,"公司名称":28,"所属指数":18,"收盘价(USD)":14,
              "当日涨跌%":12,"MA200(USD)":14,"偏离年线%":12}
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 14)
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = "A2"
    wb.save(fname)

    return df, fname, data_date, len(r1k)


def send_email(df, filepath, data_date, r1k_count=0):
    """把Excel作为附件发送，正文含完整HTML表格"""
    y, m, d = data_date.split("-")
    subject = f"SPX 500、NDX、DJCA和罗素1000指数偏离年线20%以上_{y}年{m}月{d}日"

    below_30 = len(df[df["偏离年线%"] <= -30])
    below_40 = len(df[df["偏离年线%"] <= -40])

    # 纯文本版（备用，邮件客户端不支持HTML时显示）
    text_body = f"""SPX 500 + NDX 100 + DJCA 65 + 罗素1000 每日筛选结果

数据日期：{data_date}
筛选条件：股价低于200日均线 20% 及以上
扫描范围：{'SPX500 + NDX100 + DJCA + 罗素1000' if r1k_count else '⚠ 仅 SPX500 + NDX100 + DJCA（罗素1000获取失败）'}

符合条件总数：{len(df)} 只
偏离超过 -30%：{below_30} 只
偏离超过 -40%：{below_40} 只

{df.to_string(index=False)}

（本邮件由脚本自动发送，仅供参考，不构成投资建议）
"""

    # HTML表格
    rows_html = ""
    for _, r in df.iterrows():
        dev = r["偏离年线%"]

        # 当日涨跌：涨绿跌红（美股习惯）
        dc = r.get("当日涨跌%")
        if dc is None or pd.isna(dc):
            dc_txt, dc_color = "—", "#999"
        elif dc > 0:
            dc_txt, dc_color = f"+{dc}%", "#1A7F37"
        elif dc < 0:
            dc_txt, dc_color = f"{dc}%", "#C0392B"
        else:
            dc_txt, dc_color = "0.00%", "#666"

        # 偏离越深颜色越红
        if dev <= -40:
            color = "#C0392B"; weight = "600"
        elif dev <= -30:
            color = "#E67E22"; weight = "500"
        else:
            color = "#555"; weight = "400"
        rows_html += f"""<tr>
<td style="padding:6px 10px;border-bottom:1px solid #eee;font-family:monospace;font-weight:600">{r['股票代码']}</td>
<td style="padding:6px 10px;border-bottom:1px solid #eee">{r['公司名称']}</td>
<td style="padding:6px 10px;border-bottom:1px solid #eee;font-size:12px;color:#666">{r['所属指数']}</td>
<td style="padding:6px 10px;border-bottom:1px solid #eee;text-align:right">${r['收盘价(USD)']}</td>
<td style="padding:6px 10px;border-bottom:1px solid #eee;text-align:right;color:{dc_color}">{dc_txt}</td>
<td style="padding:6px 10px;border-bottom:1px solid #eee;text-align:right;color:#888">${r['MA200(USD)']}</td>
<td style="padding:6px 10px;border-bottom:1px solid #eee;text-align:right;color:{color};font-weight:{weight}">{dev}%</td>
</tr>"""

    stale_banner = ""
    if r1k_count and R1K_CACHE_AGE_DAYS is not None and R1K_CACHE_AGE_DAYS > 400:
        stale_banner = (
            '<div style="background:#FDF3E3;border-left:4px solid #E67E22;'
            'padding:10px 14px;margin-bottom:16px;font-size:13px;color:#7A4A11">'
            f'⚠ 罗素1000名单来自 {R1K_CACHE_AGE_DAYS} 天前的缓存，已跨过一次年度重构'
            '（每年6月末），建议更新 russell1000_cache.csv。</div>'
        )

    warn_banner = stale_banner if r1k_count else (
        '<div style="background:#FDF3E3;border-left:4px solid #E67E22;'
        'padding:10px 14px;margin-bottom:16px;font-size:13px;color:#7A4A11">'
        '⚠ 本次罗素1000成分股获取失败，仅扫描了 SPX500 + NDX100 + DJCA，'
        '结果数量会明显偏少。</div>'
    )

    html_body = f"""<html><body style="font-family:-apple-system,Segoe UI,Arial,sans-serif;color:#222;line-height:1.5">
<h2 style="margin:0 0 4px;font-size:20px;font-weight:600">SPX 500 + NDX 100 + DJCA + 罗素1000 偏离年线筛选</h2>
<p style="margin:0 0 18px;color:#666;font-size:13px">数据日期 {data_date} &nbsp;·&nbsp; 条件：股价低于200日均线 20% 及以上</p>

{warn_banner}
<table style="border-collapse:collapse;margin-bottom:22px">
<tr>
  <td style="padding:10px 18px;background:#F4F6F8;border-radius:6px 0 0 6px;text-align:center">
    <div style="font-size:22px;font-weight:600">{len(df)}</div>
    <div style="font-size:11px;color:#777">符合条件</div></td>
  <td style="padding:10px 18px;background:#F4F6F8;text-align:center">
    <div style="font-size:22px;font-weight:600;color:#E67E22">{below_30}</div>
    <div style="font-size:11px;color:#777">偏离 &gt;30%</div></td>
  <td style="padding:10px 18px;background:#F4F6F8;border-radius:0 6px 6px 0;text-align:center">
    <div style="font-size:22px;font-weight:600;color:#C0392B">{below_40}</div>
    <div style="font-size:11px;color:#777">偏离 &gt;40%</div></td>
</tr></table>

<table style="border-collapse:collapse;width:100%;font-size:13px">
<thead><tr style="background:#1F3864;color:#fff">
<th style="padding:9px 10px;text-align:left">代码</th>
<th style="padding:9px 10px;text-align:left">公司名称</th>
<th style="padding:9px 10px;text-align:left">指数</th>
<th style="padding:9px 10px;text-align:right">收盘价</th>
<th style="padding:9px 10px;text-align:right">当日涨跌</th>
<th style="padding:9px 10px;text-align:right">MA200</th>
<th style="padding:9px 10px;text-align:right">偏离年线</th>
</tr></thead>
<tbody>{rows_html}</tbody>
</table>

<p style="margin-top:20px;font-size:11px;color:#999;border-top:1px solid #eee;padding-top:12px">
完整数据见附件 Excel。本邮件由脚本自动发送，仅供参考，不构成投资建议。</p>
</body></html>"""

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = SENDER_EMAIL
    msg["To"] = RECIPIENT_EMAIL
    msg.set_content(text_body)
    msg.add_alternative(html_body, subtype="html")

    with open(filepath, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(filepath),
        )

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(SENDER_EMAIL, SENDER_PASSWORD.replace(" ", ""))
        smtp.send_message(msg)

    print(f"  ✅ 邮件已发送至 {RECIPIENT_EMAIL}")
    print(f"     标题：{subject}")
    print(f"     正文含 {len(df)} 行完整表格 + Excel附件")


def main():
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    print(f"\n{'='*58}")
    print(f"  SPX500 + NDX100 + DJCA + 罗素1000 每日筛选")
    print(f"  运行时间：{now}")
    print(f"{'='*58}\n")

    if "在这里填入" in SENDER_PASSWORD:
        print("  ❌ 请先在脚本顶部填入 Gmail App Password")
        return

    df, filepath, data_date, r1k_count = run_screener()

    if df is None:
        print("  未找到符合条件的股票，不发送邮件。")
        return

    print(f"\n  找到 {len(df)} 只，Excel已生成：{os.path.basename(filepath)}")
    print(f"  正在发送邮件...")

    try:
        send_email(df, filepath, data_date, r1k_count)
    except Exception as e:
        print(f"  ❌ 邮件发送失败：{e}")
        print(f"     Excel文件仍已保存：{filepath}")
        return

    print(f"\n{'='*58}")
    print(f"  ✅ 全部完成")
    print(f"{'='*58}\n")


if __name__ == "__main__":
    main()
